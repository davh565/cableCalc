from openpyxl import Workbook
from openpyxl import load_workbook


def loadInput(input):
    wb = load_workbook(input)
    ws = wb.active
    circuitParameters = {
        "loadKw": ws['A2'].value,
        "loadAmps": ws['B2'].value,
        "cableLength": ws['C2'].value,
        "loadType": ws['D2'].value,
        "installMethod": ws['E2'].value,
        # "ambientTemp": ws['F2'].value
        # "safetyFactor": ws['G2'].value
        # "insulationType": ws['H2'].value
    }
    return circuitParameters


def saveOutput(dict, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "cableCalc output"
    ws['A1'] = "loadKw"
    ws['B1'] = "loadAmps"
    ws['C1'] = "cableLength"
    ws['D1'] = "loadType"
    ws['E1'] = "installMethod"
    ws['A2'] = dict["loadKw"]
    ws['B2'] = dict["loadAmps"]
    ws['C2'] = dict["cableLength"]
    ws['D2'] = dict["loadType"]
    ws['E2'] = dict["installMethod"]
    wb.save(output)
